"""Excel export service: generates Excel marksheets."""

from __future__ import annotations

from uuid import UUID
import io

import structlog
from shared.base_service import BaseService

logger = structlog.get_logger(__name__)


class ExcelExportService(BaseService):
    """Generates Excel files for marksheets."""

    def generate_student_marksheet_excel(self, enrollment_id: UUID) -> io.BytesIO:
        """Generate an Excel marksheet for a single student."""
        from reporting.services.marksheet_service import MarksheetService

        marksheet_svc = MarksheetService()
        marksheet = marksheet_svc.generate_student_marksheet(enrollment_id)

        buffer = io.BytesIO()
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Marksheet"

            header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            total_fill = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid")
            border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            # Header
            ws.merge_cells("A1:E1")
            ws["A1"] = f"Marksheet - {marksheet.session_name}"
            ws["A1"].font = Font(bold=True, size=14)

            # Student Info
            ws["A3"] = "Student Name"
            ws["B3"] = marksheet.student_name
            ws["A4"] = "Student ID"
            ws["B4"] = marksheet.student_id
            ws["A5"] = "Roll No"
            ws["B5"] = marksheet.roll_no
            ws["A6"] = "Class"
            ws["B6"] = f"{marksheet.class_name} - {marksheet.section_name}"

            # Subject Results Header
            row = 8
            headers = ["Subject", "Code", "Marks Obtained", "Max Marks", "Percentage", "Grade"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
                cell.border = border

            # Subject Results
            row = 9
            for subject in marksheet.subjects:
                for assessment in subject.get("assessments", []):
                    ws.cell(row=row, column=1, value=subject["subject_name"]).border = border
                    ws.cell(row=row, column=2, value=subject["subject_code"]).border = border
                    ws.cell(row=row, column=3, value=assessment["obtained_marks"]).border = border
                    ws.cell(row=row, column=4, value=assessment["full_marks"]).border = border
                    pct = round((assessment["obtained_marks"] / assessment["full_marks"]) * 100, 2) if assessment["full_marks"] > 0 else 0
                    ws.cell(row=row, column=5, value=f"{pct}%").border = border
                    row += 1

            # Totals
            row += 1
            ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
            ws.cell(row=row, column=3, value=marksheet.total_marks).font = Font(bold=True)
            ws.cell(row=row, column=4, value=marksheet.total_full).font = Font(bold=True)
            ws.cell(row=row, column=5, value=f"{marksheet.percentage}%").font = Font(bold=True)
            ws.cell(row=row, column=6, value=marksheet.overall_grade).font = Font(bold=True)
            for col in range(1, 7):
                ws.cell(row=row, column=col).fill = total_fill
                ws.cell(row=row, column=col).border = border

            # Auto-width columns
            for col in ws.columns:
                max_length = 0
                column_letter = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[column_letter].width = max_length + 2

            wb.save(buffer)
            buffer.seek(0)
            return buffer

        except ImportError:
            self.log.error("openpyxl_not_installed")
            raise RuntimeError("openpyxl is required for Excel generation")


    def generate_class_marksheet_excel(
        self,
        class_id: UUID,
        section_id: UUID,
        session_id: UUID,
    ) -> io.BytesIO:
        """Generate an Excel marksheet for all students in a class."""
        from reporting.services.marksheet_service import MarksheetService
        from enrollments.models import Enrollment

        marksheet_svc = MarksheetService()

        enrollments = Enrollment.objects.filter(
            class_field_id=class_id,
            section_id=section_id,
            session_id=session_id,
            status="active",
        ).select_related("student").order_by("roll_no")

        buffer = io.BytesIO()
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Class Marksheet"

            header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            # Header
            ws.merge_cells("A1:G1")
            ws["A1"] = f"Class Marksheet - {enrollments.first().session.name if enrollments.exists() else ''}"
            ws["A1"].font = Font(bold=True, size=14)

            # Column Headers
            row = 3
            headers = ["Roll No", "Student Name", "Student ID", "Total Marks", "Max Marks", "Percentage", "Grade"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
                cell.border = border

            # Student Data
            row = 4
            for enrollment in enrollments:
                marksheet = marksheet_svc.generate_student_marksheet(enrollment.id)
                ws.cell(row=row, column=1, value=enrollment.roll_no).border = border
                ws.cell(row=row, column=2, value=enrollment.student.name).border = border
                ws.cell(row=row, column=3, value=enrollment.student.student_id).border = border
                ws.cell(row=row, column=4, value=marksheet.total_marks).border = border
                ws.cell(row=row, column=5, value=marksheet.total_full).border = border
                ws.cell(row=row, column=6, value=f"{marksheet.percentage}%").border = border
                ws.cell(row=row, column=7, value=marksheet.overall_grade).border = border
                row += 1

            # Auto-width columns
            for col in ws.columns:
                max_length = 0
                column_letter = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[column_letter].width = max_length + 2

            wb.save(buffer)
            buffer.seek(0)
            return buffer

        except ImportError:
            self.log.error("openpyxl_not_installed")
            raise RuntimeError("openpyxl is required for Excel generation")
